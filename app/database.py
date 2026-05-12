import sqlite3
import os
import openpyxl
from datetime import datetime

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data", "goodwill.db")
DATA_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "data")


def get_db():
    conn = sqlite3.connect(DB_PATH, timeout=10)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.executescript("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code INTEGER,
            department TEXT,
            category TEXT,
            name_alternative TEXT,
            specification TEXT,
            priority_pricing INTEGER DEFAULT 0,
            own_brand INTEGER DEFAULT 0,
            description TEXT,
            packing TEXT,
            uom TEXT,
            price REAL,
            brand TEXT,
            sub_group TEXT,
            model_no TEXT,
            tax_perc REAL DEFAULT 0,
            stock_available INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_code INTEGER UNIQUE,
            customer_name TEXT,
            location TEXT,
            payment_terms TEXT DEFAULT 'cash',
            order_status TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        );

        CREATE TABLE IF NOT EXISTS customer_contacts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            contact_person TEXT,
            phone_number TEXT,
            is_active INTEGER DEFAULT 1,
            added_by TEXT DEFAULT 'system',
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );

        CREATE TABLE IF NOT EXISTS customer_call_schedule (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER UNIQUE,
            calling_time TEXT,
            rescheduled_time TEXT,
            last_called_at TEXT,
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );

        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_number TEXT UNIQUE,
            customer_id INTEGER,
            caller_phone TEXT,
            caller_name TEXT,
            total_amount REAL DEFAULT 0,
            currency TEXT DEFAULT 'QR',
            payment_terms TEXT,
            delivery_schedule TEXT,
            status TEXT DEFAULT 'pending',
            notes TEXT,
            wallet_discount REAL DEFAULT 0,
            wallet_upsell REAL DEFAULT 0,
            wallet_balance REAL DEFAULT 0,
            conversation_id TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (customer_id) REFERENCES customers(id)
        );

        CREATE TABLE IF NOT EXISTS order_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            order_id INTEGER,
            product_id INTEGER,
            product_description TEXT,
            packing TEXT,
            uom TEXT,
            quantity INTEGER DEFAULT 1,
            unit_price REAL,
            total_price REAL,
            discount REAL DEFAULT 0,
            FOREIGN KEY (order_id) REFERENCES orders(id),
            FOREIGN KEY (product_id) REFERENCES products(id)
        );

        CREATE TABLE IF NOT EXISTS call_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            conversation_id TEXT,
            caller_phone TEXT,
            customer_id INTEGER,
            customer_name TEXT,
            direction TEXT DEFAULT 'inbound',
            language TEXT,
            duration_seconds INTEGER,
            status TEXT DEFAULT 'completed',
            summary TEXT,
            transcript TEXT,
            order_id INTEGER,
            has_complaint INTEGER DEFAULT 0,
            complaint_details TEXT,
            callback_required INTEGER DEFAULT 0,
            callback_reason TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (order_id) REFERENCES orders(id)
        );

        CREATE TABLE IF NOT EXISTS complaints (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            call_log_id INTEGER,
            category TEXT,
            description TEXT,
            status TEXT DEFAULT 'open',
            assigned_to TEXT,
            resolved_at TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (call_log_id) REFERENCES call_logs(id)
        );

        CREATE TABLE IF NOT EXISTS wallet_transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            customer_id INTEGER,
            order_id INTEGER,
            transaction_type TEXT,
            amount REAL,
            description TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (customer_id) REFERENCES customers(id),
            FOREIGN KEY (order_id) REFERENCES orders(id)
        );

        CREATE INDEX IF NOT EXISTS idx_products_department ON products(department);
        CREATE INDEX IF NOT EXISTS idx_products_category ON products(category);
        CREATE INDEX IF NOT EXISTS idx_products_code ON products(code);
        CREATE INDEX IF NOT EXISTS idx_products_own_brand ON products(own_brand);
        CREATE INDEX IF NOT EXISTS idx_customer_contacts_phone ON customer_contacts(phone_number);
        CREATE INDEX IF NOT EXISTS idx_orders_customer ON orders(customer_id);
        CREATE INDEX IF NOT EXISTS idx_orders_status ON orders(status);
        CREATE INDEX IF NOT EXISTS idx_call_logs_customer ON call_logs(customer_id);
        CREATE INDEX IF NOT EXISTS idx_wallet_customer ON wallet_transactions(customer_id);
    """)

    try:
        c.execute("ALTER TABLE customers ADD COLUMN delivery_preference TEXT DEFAULT 'morning'")
    except Exception:
        pass

    conn.commit()
    conn.close()


def load_products_from_excel():
    filepath = os.path.join(DATA_DIR, "ai worksheet.xlsx")
    if not os.path.exists(filepath):
        print(f"Product file not found: {filepath}")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM products")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_map = {h: i for i, h in enumerate(headers) if h}

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[col_map.get("Code", 0)]
        if code is None:
            continue

        department = row[col_map.get("Department", 1)] or ""
        category = row[col_map.get("Category", 2)] or ""
        name_alt = row[col_map.get("name alternative", 3)] or ""
        spec = row[col_map.get("specification", 4)] or ""
        priority = 1 if row[col_map.get("priority pricing", 5)] else 0
        own_brand = 1 if row[col_map.get("own brand", 6)] else 0
        desc = row[col_map.get("Description", 7)] or ""
        packing = row[col_map.get("Packing", 8)] or ""
        uom = row[col_map.get("UOM", 9)] or ""
        price = row[col_map.get("Price", 10)] or 0
        brand = row[col_map.get("Brand (Sub Category)", 12)] or ""
        sub_group = row[col_map.get("Sub Group", 13)] or ""
        model_no = row[col_map.get("Model No", 14)] or ""
        tax_perc = row[col_map.get("TaxPerc", 15)] or 0

        c.execute("""
            INSERT INTO products (code, department, category, name_alternative, specification,
                priority_pricing, own_brand, description, packing, uom, price,
                brand, sub_group, model_no, tax_perc)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (code, department.lower().strip(), category.lower().strip(), name_alt, spec,
              priority, own_brand, desc, packing, uom.lower().strip(), price,
              brand, sub_group, model_no, tax_perc))
        count += 1

    conn.commit()
    conn.close()
    print(f"Loaded {count} products")


def load_customers_from_excel():
    filepath = os.path.join(DATA_DIR, "ai customer list.xlsx")
    if not os.path.exists(filepath):
        print(f"Customer file not found: {filepath}")
        return

    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM customer_contacts")
    c.execute("DELETE FROM customer_call_schedule")
    c.execute("DELETE FROM customers")

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        if code is None:
            continue

        name = row[1] or ""
        location = row[2] or ""
        payment = row[12] or "cash"
        order_status = row[11] or ""

        c.execute("""
            INSERT INTO customers (customer_code, customer_name, location, payment_terms, order_status)
            VALUES (?, ?, ?, ?, ?)
        """, (code, name.lower().strip(), location.lower().strip(),
              payment.lower().strip(), order_status))
        customer_id = c.lastrowid

        for i in range(3):
            contact = row[3 + i * 2]
            phone = row[4 + i * 2]
            if contact and phone:
                phone_str = str(int(phone)) if isinstance(phone, (int, float)) else str(phone)
                c.execute("""
                    INSERT INTO customer_contacts (customer_id, contact_person, phone_number)
                    VALUES (?, ?, ?)
                """, (customer_id, str(contact).lower().strip(), phone_str))

        calling_time = row[9] or ""
        rescheduled = row[10] or ""
        c.execute("""
            INSERT INTO customer_call_schedule (customer_id, calling_time, rescheduled_time)
            VALUES (?, ?, ?)
        """, (customer_id, str(calling_time), str(rescheduled) if rescheduled else None))

        count += 1

    conn.commit()
    conn.close()
    print(f"Loaded {count} customers")


def seed_database():
    init_db()
    load_products_from_excel()
    load_customers_from_excel()
    print("Database seeded successfully")


if __name__ == "__main__":
    seed_database()
